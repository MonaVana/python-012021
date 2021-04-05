
import pandas
praha = pandas.read_csv("zam_praha.csv")
plzen = pandas.read_csv("zam_plzeň.csv")
liberec = pandas.read_csv("zam_liberec.csv")
platy = pandas.read_csv("platy_2021_02.csv")

praha["mesto"] = "Praha"
plzen["mesto"] = "Plzeň"
liberec["mesto"] = "Liberec"

zamestnanci = pandas.concat([praha, plzen, liberec], ignore_index=True)

platy_zamestnancu = pandas.merge(zamestnanci, platy, on=["cislo_zamestnance"], how="left")

platy_zamestnancu.to_excel("platy zaměstnanců.xlsx", index=False)

platy_zamestnancu = pandas.read_excel("platy zaměstnanců.xlsx")

#doplněk

from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
radek = 1
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
    bunka = ws1.cell(radek, sloupec)
    bunka.value = predmet
    sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
    bunka.fill = sediva_barva
    sloupec += 1
  radek += 1

wb.save(filename="rozvrh_hodin.xlsx")


